import datetime as dt

import openpyxl
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import AttendanceDaily, Parameter, SchoolCalendar, Student


def _headers(ws):
    return [
        (c.value or "").strip() if isinstance(c.value, str) else c.value
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]


def _rows_as_dicts(ws):
    headers = _headers(ws)
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        record = dict(zip(headers, values))
        yield record


def _to_time_str(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.time):
        return value.strftime("%H:%M")
    if isinstance(value, dt.datetime):
        return value.strftime("%H:%M")
    if isinstance(value, str):
        return value.strip() or None
    return None


def _to_date(value) -> dt.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str) and value.strip():
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return dt.datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _to_bool_ya_tidak(value) -> bool:
    return isinstance(value, str) and value.strip().lower() == "ya"


def _to_int(value) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


async def import_workbook(db: AsyncSession, path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    counts = {"students": 0, "attendance": 0, "calendar": 0, "parameters": 0}

    if "MASTER_SISWA" in wb.sheetnames:
        rows = []
        for r in _rows_as_dicts(wb["MASTER_SISWA"]):
            nis = r.get("NIS")
            if not nis or not str(nis).strip():
                continue
            rows.append(
                {
                    "nis": str(nis).strip(),
                    "nisn": str(r.get("NISN") or "").strip() or None,
                    "nama": str(r.get("Nama Siswa") or "").strip(),
                    "jenis_kelamin": (str(r.get("Jenis Kelamin") or "").strip() or None),
                    "tanggal_lahir": _to_date(r.get("Tanggal Lahir")),
                    "jenjang": str(r.get("Jenjang") or "").strip() or None,
                    "kelas": str(r.get("Kelas") or "").strip() or None,
                    "wali_kelas": str(r.get("Wali Kelas") or "").strip() or None,
                    "tahun_ajaran": str(r.get("Tahun Ajaran") or "").strip() or None,
                    "status_siswa": str(r.get("Status Siswa") or "Aktif").strip(),
                }
            )
        if rows:
            stmt = pg_insert(Student).values(rows)
            update_cols = {c.name: getattr(stmt.excluded, c.name) for c in Student.__table__.columns if c.name != "nis"}
            stmt = stmt.on_conflict_do_update(index_elements=[Student.nis], set_=update_cols)
            await db.execute(stmt)
            counts["students"] = len(rows)

    if "ABSENSI_HARIAN" in wb.sheetnames:
        rows = []
        for r in _rows_as_dicts(wb["ABSENSI_HARIAN"]):
            nis = r.get("NIS")
            tanggal = _to_date(r.get("Tanggal"))
            if not nis or not str(nis).strip() or tanggal is None:
                continue
            rows.append(
                {
                    "tanggal": tanggal,
                    "nis": str(nis).strip(),
                    "nama_siswa": str(r.get("Nama Siswa") or "").strip(),
                    "jenjang": str(r.get("Jenjang") or "").strip() or None,
                    "kelas": str(r.get("Kelas") or "").strip() or None,
                    "wali_kelas": str(r.get("Wali Kelas") or "").strip() or None,
                    "jam_masuk": _to_time_str(r.get("Jam Masuk")),
                    "jam_pulang": _to_time_str(r.get("Jam Pulang")),
                    "status_kehadiran": str(r.get("Status Kehadiran") or "").strip(),
                    "keterangan": str(r.get("Keterangan") or "").strip() or None,
                    "menit_terlambat": _to_int(r.get("Menit Terlambat")),
                    "sudah_absen_masuk": _to_bool_ya_tidak(r.get("Sudah Absen Masuk")),
                    "belum_absen_masuk": _to_bool_ya_tidak(r.get("Belum Absen Masuk")),
                    "sudah_absen_pulang": _to_bool_ya_tidak(r.get("Sudah Absen Pulang")),
                    "belum_absen_pulang": _to_bool_ya_tidak(r.get("Belum Absen Pulang")),
                    "durasi_menit": _to_int(r.get("Durasi di Sekolah (Menit)")) or None,
                    "sumber_data": str(r.get("Sumber Data") or "").strip() or None,
                }
            )
        if rows:
            stmt = pg_insert(AttendanceDaily).values(rows)
            update_cols = {
                c.name: getattr(stmt.excluded, c.name)
                for c in AttendanceDaily.__table__.columns
                if c.name not in ("id", "tanggal", "nis")
            }
            stmt = stmt.on_conflict_do_update(
                index_elements=[AttendanceDaily.tanggal, AttendanceDaily.nis], set_=update_cols
            )
            await db.execute(stmt)
            counts["attendance"] = len(rows)

    if "KALENDER" in wb.sheetnames:
        rows = []
        for r in _rows_as_dicts(wb["KALENDER"]):
            tanggal = _to_date(r.get("Tanggal"))
            if tanggal is None:
                continue
            rows.append(
                {
                    "tanggal": tanggal,
                    "hari": str(r.get("Hari") or "").strip() or None,
                    "bulan": str(r.get("Bulan") or "").strip() or None,
                    "tahun": str(r.get("Tahun") or "").strip() or None,
                    "semester": str(r.get("Semester") or "").strip() or None,
                    "tahun_ajaran": str(r.get("Tahun Ajaran") or "").strip() or None,
                    "jenis_hari": str(r.get("Jenis Hari") or "").strip() or None,
                    "hari_sekolah": str(r.get("Hari Sekolah") or "").strip() or None,
                }
            )
        if rows:
            stmt = pg_insert(SchoolCalendar).values(rows)
            update_cols = {
                c.name: getattr(stmt.excluded, c.name)
                for c in SchoolCalendar.__table__.columns
                if c.name != "tanggal"
            }
            stmt = stmt.on_conflict_do_update(index_elements=[SchoolCalendar.tanggal], set_=update_cols)
            await db.execute(stmt)
            counts["calendar"] = len(rows)

    if "PARAMETER" in wb.sheetnames:
        rows = []
        for r in _rows_as_dicts(wb["PARAMETER"]):
            key = r.get("Parameter")
            if not key or not str(key).strip():
                continue
            rows.append(
                {
                    "key": str(key).strip(),
                    "value": str(r.get("Nilai") or "").strip(),
                    "keterangan": str(r.get("Keterangan") or "").strip() or None,
                }
            )
        if rows:
            stmt = pg_insert(Parameter).values(rows)
            update_cols = {"value": stmt.excluded.value, "keterangan": stmt.excluded.keterangan}
            stmt = stmt.on_conflict_do_update(index_elements=[Parameter.key], set_=update_cols)
            await db.execute(stmt)
            counts["parameters"] = len(rows)

    await db.commit()
    return counts
